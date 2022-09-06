from openpyxl import load_workbook
import string
from test_scripts.voc_generator import make_random_voc


def write_n_rows(ws, last_line, image_info):
    alphabet = string.ascii_uppercase
    for i in range(len(image_info)):
        ws[f'{alphabet[i]}{last_line + 1}'] = image_info[alphabet[i]]


def write_to_file(path_to_file, image_info, last_line, report_date):
    wb = load_workbook(path_to_file)
    ws = wb[report_date]
    write_n_rows(ws, last_line + 1, image_info)
    wb.save(path_to_file)


def write_xlsx_single_sheet(path_to_file, image_info):
    wb = load_workbook(path_to_file)
    ws = wb.active
    last_line = ws.max_row
    write_n_rows(ws, last_line, image_info)
    wb.save(path_to_file)


path_to_file = '/Users/evgeniy/Documents/Kommersant/My_report_from_0107/report_file_July_test.xlsx'

image_info = make_random_voc(5)
write_xlsx_single_sheet(path_to_file, image_info)
