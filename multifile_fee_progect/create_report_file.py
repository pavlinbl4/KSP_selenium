from openpyxl import load_workbook, Workbook
from pathlib import Path
from kommersant_dates import KommersantDates
import os


def create_report_file(report_date):
    kd = KommersantDates()
    month_name = kd.previous_month_name
    report_folder = f'{Path.home()}/Documents/Kommersant/My_report_from_0107'
    report_file_name = f"report_file_{month_name}_test.xlsx"
    path_to_file = f'{report_folder}/{report_file_name}'

    if os.path.exists(path_to_file):
        wb = load_workbook(path_to_file)  # файл есть и открываю его
        ws = wb.create_sheet(report_date)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файла еще нет
        ws = wb.active  # если файла еще нет
        ws.title = report_date  # если файла еще нет

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25  # задаю ширину колонки
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 100

    ws['A1'] = 'number'
    ws['B1'] = 'KSP_id'  # create columns names
    ws['C1'] = 'date of publication'
    ws['D1'] = 'publication'
    ws['E1'] = 'material'

    wb.save(path_to_file)
    return path_to_file
