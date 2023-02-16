from openpyxl import load_workbook
from openpyxl import Workbook
import os
from datetime import datetime
from pathlib import Path


def make_documents_subfolder(name):  # create folder "name" in User/Documents folder
    (Path.home() / "Documents" / f"{name}").mkdir(parents=True, exist_ok=True)
    return Path.home() / "Documents" / f"{name}"


def create_report_file(report_folder, file_name):
    report_date = datetime.today().strftime("%Y-%m-%d")
    report_file_name = f"{report_date}_{file_name}.xlsx"
    path_to_file = f'{report_folder}/{report_file_name}'

    if os.path.exists(path_to_file):
        wb = load_workbook(path_to_file)  # файл есть и открываю его
        ws = wb.create_sheet(report_date)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файа еще нет
        ws.title = report_date  # если файда еще нет

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20  # задаю ширину колонки
    ws.column_dimensions['D'].width = 20

    ws['A1'] = 'KSP_id'  # create columns names
    ws['B1'] = 'date of publication'
    ws['C1'] = 'publication'
    ws['D1'] = 'material'

    wb.save(path_to_file)


def write_to_file(path_to_file, line_number, report_date):
    wb = load_workbook(path_to_file)
    ws = wb[report_date]
    ws[f'A{line_number}'] = line_number
    ws[f'B{line_number}'] = line_number
    ws[f'C{line_number}'] = 'publication'
    ws[f'D{line_number}'] = 'material'

    wb.save(path_to_file)


create_report_file(make_documents_subfolder(input("Enter the folder name\n\n")), input("Enter report name\n\n"))

if __name__ == '__main__':
    create_report_file(make_documents_subfolder(input("Enter the folder name\n\n")), input("Enter report name\n\n"))
