"""
script makes report for previous month and save it to file
but in  some sheets (18.07.2022) I found blank lines
"""
from calendar import monthrange
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from datetime import datetime, timedelta
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import os
from dotenv import load_dotenv
from bs4 import BeautifulSoup
import re
from selenium.webdriver.common.alert import Alert
from openpyxl import load_workbook
from openpyxl import Workbook

load_dotenv()
login = os.environ.get('login')
password = os.environ.get('password')
first_loggin = os.environ.get('first_loggin')

report_web_link = 'https://image.kommersant.ru/photo/archive/pubhistory.asp?ID='

image_info = {}
images_voc = {}


def days_in_previous_month():  # days in last month
    current_month = datetime.today().month
    if current_month != 1:
        previous_month = current_month - 1
        report_year = datetime.today().year
    else:
        previous_month = 12
        report_year = datetime.today().year - 1
    return monthrange(report_year, previous_month)[1]


def create_report_file(report_date):
    report_month = (datetime.today().replace(day=1) - timedelta(days=1)).strftime('%B')
    report_folder = '/Volumes/big4photo/Documents/Kommersant/My_report_from_0107'
    report_file_name = f"report_file_{report_month}.xlsx"
    path = f'{report_folder}/{report_file_name}'

    if os.path.exists(path):
        wb = load_workbook(path)  # файл есть и открываю его
        ws = wb.create_sheet(report_date)  # добавляю новую таблицу
    else:
        wb = Workbook()  # если файда еще нет
        ws = wb.active  # если файа еще нет
        ws.title = report_date  # если файда еще нет

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25  # задаю ширину колонки
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 100

    ws['A1'] = 'number'
    ws['B1'] = 'KSP_id'  # create columns names
    ws['C1'] = 'date of publication'
    ws['D1'] = 'publication'
    ws['E1'] = 'material'

    wb.save(path)
    return path


def write_to_file(line_number, report_date):
    wb = load_workbook(path_to_file)
    ws = wb[report_date]
    if len(image_info) > 0:
        ws[f'A{line_number + 2}'] = image_info['A']
        ws[f'B{line_number + 2}'] = image_info['B']
        ws[f'C{line_number + 2}'] = image_info['C']
        ws[f'D{line_number + 2}'] = image_info['D']
        ws[f'E{line_number + 2}'] = image_info['E']

    wb.save(path_to_file)


def save_html():
    with open(f'scrap_report_{k}.html', 'w') as saved_file:
        saved_file.write(html)


def publication_info():
    report_link = f'{report_web_link}{k}#web'
    browser.get(report_link)
    print(f'{report_link = }')
    report_html = browser.page_source
    soup = BeautifulSoup(report_html, 'lxml')
    all_publications = soup.find(id='Table1').find('tbody').find_all('tr')
    for i in range(len(all_publications)):
        try:
            if report_day in all_publications[i].find_all('td')[8].text:  # date of upload
                print(f"{count} - {soup.find('h3').text}")
                image_info['A'] = count
                image_info['B'] = soup.find('h3').text[16:]
                print(f"date of publication - {all_publications[i].find_all('td')[2].text}")  # date of publication
                image_info["C"] = all_publications[i].find_all('td')[2].text
                print(f"publication - {all_publications[i].find_all('td')[3].text}")  # publication
                image_info["D"] = all_publications[i].find_all('td')[3].text

                print(f"material - {all_publications[i].find_all('td')[5].text}")  # material
                image_info["E"] = all_publications[i].find_all('td')[5].text
        finally:
            continue

    print()
    return image_info


def make_images_voc():
    for _ in images_links:
        time.sleep(3)
        ksp_id = re.findall(r'(?<=photocode=)\w{16}', str(_))[0]
        regex = r'(?<=photoid=)\d{7}(?=\")'
        photoid = re.findall(regex, str(_))[0]
        images_voc[photoid] = ksp_id
    return images_voc


def get_image_links():
    soup = BeautifulSoup(html, 'lxml')
    return soup.find_all('table')[9].find('tbody').find_all(title="Добавить кадрировку")


def setting_chrome_options():
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # фоновый режим
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # невидимость автоматизации
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
    return chrome_options


def published_images_amount():
    try:
        images_amount = \
            browser.find_element(By.XPATH, '/html/body/table[3]/tbody/tr[1]/td[2]/table/tbody/tr[2]/td/b[1]').text

        print(f'used images {images_amount}\n')
    except Exception as ex:
        print(ex)
        print('no published images in this day')


def autorization():  # авторизация на главной странице
    browser.get(first_loggin)
    login_input = browser.find_element(By.ID, "login")
    login_input.send_keys(login)
    password_input = browser.find_element(By.ID, "password")
    password_input.send_keys(password)
    browser.find_element(By.CSS_SELECTOR, ".system input.but").click()
    browser.find_element(By.CSS_SELECTOR, '#au').send_keys('Евгений Павленко')


def select_today_published_images():
    autorization()
    time.sleep(1)

    try:
        alert = Alert(browser)
        alert.accept()
        time.sleep(2)
    finally:
        select = Select(browser.find_element(By.NAME, 'ps'))
        select.select_by_value('50')

        select = Select(browser.find_element(By.ID, "dt"))  # select "засыла"
        select.select_by_value("3")

        data_input = browser.find_element(By.ID, "since")
        data_input.clear()
        data_input.send_keys(report_day)

        data_input = browser.find_element(By.ID, "till")
        data_input.clear()
        data_input.send_keys(report_day)

        browser.find_element(By.CSS_SELECTOR, '#searchbtn').click()

        try:
            alert = Alert(browser)
            alert.accept()
            time.sleep(2)
        finally:

            return browser.page_source


if __name__ == '__main__':
    days_in_month = days_in_previous_month()
    start_day = (datetime.today().replace(day=1) - timedelta(days=1))
    for x in range(days_in_month - 1, -1, -1):
        time.sleep(10)
        browser = webdriver.Chrome(options=setting_chrome_options())

        report_day = (start_day - timedelta(days=x)).strftime("%d.%m.%Y")
        print(f'{report_day = }')

        path_to_file = create_report_file(report_day)
        html = select_today_published_images()
        published_images_amount()
        images_links = get_image_links()
        images_voc = make_images_voc()

        count = 0
        for k in images_voc:
            count += 1
            image_info = publication_info()
            write_to_file(count, report_day)

        browser.close()
        browser.quit()
