"""
добавляет в файл
/Volumes/big4photo/Documents/Kommersant/publications/shoot_publications.xlsx
данные о публикации снимков за заданный период, при повторном запросе одного и того же месяца
создает новую вкладку, доработать задание ширины столбцов
"""
from selenium import webdriver
from datetime import datetime
import time
from openpyxl import load_workbook
import calendar
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
import os
from dotenv import load_dotenv

load_dotenv()
login = os.environ.get('login')
password = os.environ.get('password')
first_loggin = os.environ.get('first_loggin')
print(login)

current_year = datetime.now().year



month_number = int(input("введите номер месяца\n"))
day_in_month = calendar.mdays[month_number]


options = webdriver.ChromeOptions()
options.add_argument("user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
browser = webdriver.Chrome(options=options)

try:
    browser.get(first_loggin)
    login_input = browser.find_element(By.ID, "login")
    login_input.send_keys(login)

    password_input = browser.find_element(By.ID, "password")
    password_input.send_keys(password)

    browser.find_element(By.CSS_SELECTOR, ".system input.but").click()
    author_input = browser.find_element(By.CSS_SELECTOR, "input#au")
    author_input.send_keys("Евгений Павленко")

    select = Select(browser.find_element(By.ID, "dt"))
    select.select_by_value("3")

    # записываем в xlxs
    book = load_workbook("/Volumes/big4photo/Documents/Kommersant/publications/shoot_publications.xlsx")  # открываю существующий файл для дозаписи
    # sheet = book.active
    ws_month_number = book.create_sheet(f"{current_year}_{calendar.month_name[month_number]}", 0)

    publications = {}
    for i in range(1, day_in_month + 1):
        day = str(i) + f".{month_number}.{current_year}"   # автоматизировать год
        publications[day] = []

        # вставляем сегодняшнюю дату в меню поиска
        data_input = browser.find_element(By.ID, "since")
        data_input.clear()
        data_input.send_keys(day)

        data_input = browser.find_element(By.ID, "till")
        data_input.clear()
        data_input.send_keys(day)

        browser.find_element(By.ID, "searchbtn").click()

        images_id = browser.find_elements(By.CLASS_NAME,"code")  # нахожу коды всех снимков на странице



        ws_month_number.cell(row=1, column=1).value = "Дата"
        ws_month_number.cell(row=3, column=1).value = "photo id"
        count = 0
        for id in images_id:
            if "KSP" in id.text:
                publications[day].append(id.text[4:20])
                count += 1
                ws_month_number.cell(row=2 + count, column=1 + i).value = id.text[4:20]
        ws_month_number.cell(row=1, column=1 + i).value = day


        book.save(f"/Volumes/big4photo/Documents/Kommersant/publications/shoot_publications.xlsx")
        book.close()
    print(publications)

    time.sleep(10)
    browser.close()
    browser.quit()
except Exception as ex:
    print('mistake')
    print(ex)
    browser.close()
    browser.quit()
